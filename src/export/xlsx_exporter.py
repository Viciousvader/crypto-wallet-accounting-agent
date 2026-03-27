from __future__ import annotations

from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

COLUMN_HEADERS = [
    "row_id",
    "timestamp_utc",
    "tx_hash",
    "chain",
    "source_type",
    "event_type",
    "classification_confidence",
    "review_flag",
    "review_reason",
    "pricing_status",
    "price_source",
    "direction",
    "asset_symbol",
    "amount",
    "price_usd",
    "usd_value",
    "fee_asset",
    "fee_amount",
    "counterparty",
    "from_address",
    "to_address",
    "contract_address",
    "method_label",
    "status",
]

SUMMARY_HEADERS = ["metric", "value"]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
LOW_CONF_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F8FBFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

AMOUNT_HEADERS = {"amount", "fee_amount", "price_usd", "usd_value"}
WRAP_HEADERS = {
    "tx_hash",
    "review_reason",
    "counterparty",
    "from_address",
    "to_address",
    "contract_address",
    "method_label",
}
CENTER_HEADERS = {
    "chain",
    "source_type",
    "event_type",
    "classification_confidence",
    "review_flag",
    "pricing_status",
    "price_source",
    "direction",
    "asset_symbol",
    "fee_asset",
    "status",
}
SUMMARY_SECTION_BREAKS = {
    "wallet_address",
    "classified_record_count",
    "summary.asset_symbol_counts",
    "summary.event_counts",
    "summary.confidence_counts",
    "summary.review_counts",
    "summary.review_reason_counts",
    "summary.pricing_status_counts",
    "summary.pricing_review_reason_counts",
    "asset_symbol_count.ETH",
}
ETHERSCAN_TX_BASE = "https://etherscan.io/tx/"


def export_classified_history_to_workbook(
    classified_payload: dict[str, Any],
    output_path: Path,
) -> dict[str, Any]:
    records = [record for record in classified_payload.get("records", []) if isinstance(record, dict)]
    summary = classified_payload.get("summary", {}) if isinstance(classified_payload, dict) else {}

    workbook = Workbook()
    all_sheet = workbook.active
    all_sheet.title = "All Transactions"
    review_sheet = workbook.create_sheet("Needs Review")
    summary_sheet = workbook.create_sheet("Summary")

    _write_records_sheet(all_sheet, records)
    _write_records_sheet(review_sheet, [record for record in records if record.get("review_flag")], highlight_review=True)
    _write_summary_sheet(summary_sheet, classified_payload, summary)
    _style_workbook_tabs(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "output_path": str(output_path),
        "sheet_counts": {
            "all_transactions": len(records),
            "needs_review": sum(1 for record in records if record.get("review_flag")),
            "summary_rows": summary_sheet.max_row - 1,
        },
    }


def _write_records_sheet(sheet: Worksheet, records: list[dict[str, Any]], highlight_review: bool = False) -> None:
    sheet.append(COLUMN_HEADERS)
    _style_header_row(sheet)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_HEADERS))}1"

    for row_index, record in enumerate(records, start=2):
        row_values = [_coerce_cell_value(record.get(header)) for header in COLUMN_HEADERS]
        sheet.append(row_values)
        _style_record_row(sheet, row_index, record, highlight_review=highlight_review)

    sheet.freeze_panes = "A2"
    _apply_sheet_view(sheet)
    _autosize_columns(sheet)
    _tune_record_sheet_layout(sheet)


def _write_summary_sheet(sheet: Worksheet, classified_payload: dict[str, Any], summary: dict[str, Any]) -> None:
    sheet.append(SUMMARY_HEADERS)
    _style_header_row(sheet)

    wallet_address = str(classified_payload.get("wallet_address", "") or "")
    chain = str(classified_payload.get("chain", "") or "")
    records = [record for record in classified_payload.get("records", []) if isinstance(record, dict)]

    summary_rows: list[tuple[str, Any]] = [
        ("wallet_address", wallet_address),
        ("chain", chain),
        ("classified_record_count", len(records)),
        ("summary.asset_symbol_counts", ""),
    ]

    for key, value in sorted(summary.items()):
        if isinstance(value, dict):
            summary_rows.append((f"summary.{key}", ""))
            for nested_key, nested_value in sorted(value.items()):
                summary_rows.append((f"{key}.{nested_key}", nested_value))
        else:
            summary_rows.append((key, value))

    asset_counts = Counter(str(record.get("asset_symbol", "") or "") for record in records if record.get("asset_symbol"))
    for asset_symbol, count in sorted(asset_counts.items()):
        summary_rows.append((f"asset_symbol_count.{asset_symbol}", count))

    for metric, value in summary_rows:
        sheet.append([metric, _coerce_cell_value(value)])

    _style_summary_sheet(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:B1"
    _apply_sheet_view(sheet)
    _autosize_columns(sheet)
    sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width or 0, 30)
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 18)


def _style_header_row(sheet: Worksheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_record_row(sheet: Worksheet, row_index: int, record: dict[str, Any], highlight_review: bool = False) -> None:
    review_flag = bool(record.get("review_flag"))
    low_conf = str(record.get("classification_confidence", "") or "").upper() == "LOW"
    failed = str(record.get("status", "") or "").lower() == "failed"

    fill = None
    if review_flag:
        fill = REVIEW_FILL
    elif low_conf:
        fill = LOW_CONF_FILL
    elif failed:
        fill = FAIL_FILL
    elif row_index % 2 == 0:
        fill = ALT_ROW_FILL

    for col_index, header in enumerate(COLUMN_HEADERS, start=1):
        cell = sheet.cell(row=row_index, column=col_index)
        cell.border = THIN_BORDER
        if fill is not None:
            cell.fill = fill

        if header in AMOUNT_HEADERS:
            cell.alignment = Alignment(horizontal="right", vertical="top")
            cell.number_format = "0.############################"
        elif header in WRAP_HEADERS:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        elif header in CENTER_HEADERS:
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="top")

    tx_hash = str(record.get("tx_hash", "") or "")
    if tx_hash:
        tx_cell = sheet.cell(row=row_index, column=COLUMN_HEADERS.index("tx_hash") + 1)
        tx_cell.hyperlink = f"{ETHERSCAN_TX_BASE}{tx_hash}"
        tx_cell.style = "Hyperlink"

    if highlight_review and not review_flag:
        for col_index in range(1, len(COLUMN_HEADERS) + 1):
            sheet.cell(row=row_index, column=col_index).fill = REVIEW_FILL


def _style_summary_sheet(sheet: Worksheet) -> None:
    for row_index in range(2, sheet.max_row + 1):
        metric_cell = sheet.cell(row=row_index, column=1)
        value_cell = sheet.cell(row=row_index, column=2)

        metric = str(metric_cell.value or "")
        metric_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER

        if metric.startswith("summary.") and not value_cell.value:
            metric_cell.fill = SUBHEADER_FILL
            value_cell.fill = SUBHEADER_FILL
            metric_cell.font = Font(bold=True)
            value_cell.font = Font(bold=True)
        elif row_index % 2 == 0:
            metric_cell.fill = ALT_ROW_FILL
            value_cell.fill = ALT_ROW_FILL

        if metric in SUMMARY_SECTION_BREAKS:
            metric_cell.font = Font(bold=True)

        metric_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _apply_sheet_view(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = True
    sheet.sheet_view.zoomScale = 90


def _tune_record_sheet_layout(sheet: Worksheet) -> None:
    header_to_col = {header: index + 1 for index, header in enumerate(COLUMN_HEADERS)}

    fixed_widths = {
        "row_id": 10,
        "timestamp_utc": 22,
        "tx_hash": 20,
        "chain": 12,
        "source_type": 20,
        "event_type": 16,
        "classification_confidence": 12,
        "review_flag": 10,
        "review_reason": 24,
        "pricing_status": 14,
        "price_source": 12,
        "direction": 10,
        "asset_symbol": 12,
        "amount": 18,
        "price_usd": 14,
        "usd_value": 16,
        "fee_asset": 10,
        "fee_amount": 14,
        "counterparty": 20,
        "from_address": 20,
        "to_address": 20,
        "contract_address": 20,
        "method_label": 24,
        "status": 10,
    }

    for header, width in fixed_widths.items():
        col_letter = get_column_letter(header_to_col[header])
        sheet.column_dimensions[col_letter].width = width

    for row_index in range(2, sheet.max_row + 1):
        review_reason = str(sheet.cell(row=row_index, column=header_to_col["review_reason"]).value or "")
        method_label = str(sheet.cell(row=row_index, column=header_to_col["method_label"]).value or "")
        from_address = str(sheet.cell(row=row_index, column=header_to_col["from_address"]).value or "")
        to_address = str(sheet.cell(row=row_index, column=header_to_col["to_address"]).value or "")
        max_len = max(len(review_reason), len(method_label), len(from_address), len(to_address))
        sheet.row_dimensions[row_index].height = 30 if max_len > 24 else 18


def _style_workbook_tabs(workbook: Workbook) -> None:
    workbook["All Transactions"].sheet_properties.tabColor = "1F4E78"
    workbook["Needs Review"].sheet_properties.tabColor = "C27C0E"
    workbook["Summary"].sheet_properties.tabColor = "38761D"


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _coerce_cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return value
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, Decimal):
        return format(value, "f")

    text = str(value)
    if _looks_like_decimal(text):
        return text
    return text


def _looks_like_decimal(text: str) -> bool:
    try:
        Decimal(text)
        return True
    except (InvalidOperation, ValueError):
        return False
